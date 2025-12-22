import openpyxl
import re

# Load the workbook
wb = openpyxl.load_workbook('HPV_25_26-zminy-hruden.xlsx')

print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}\n")

# Analyze each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")

    # Show first 15 rows to understand structure
    print("\nFirst 15 rows:")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        # Filter out empty rows
        if any(cell is not None for cell in row):
            print(f"Row {i}: {row}")

    print("\n" + "-"*60)

    # Try to find addresses (streets)
    address_pattern = re.compile(r'(вул\.|пров\.|просп\.|бульв\.)', re.IGNORECASE)
    found_addresses = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        for cell_idx, cell in enumerate(row):
            if cell and isinstance(cell, str):
                if address_pattern.search(cell):
                    found_addresses.append((row_idx, cell_idx, cell))

    print(f"\nFound {len(found_addresses)} potential addresses")
    if found_addresses:
        print("Sample addresses:")
        for row, col, addr in found_addresses[:10]:
            print(f"  Row {row}, Col {col}: {addr}")

wb.close()

print("\n" + "="*60)
print("SUMMARY")
print("="*60)
