import openpyxl
import re
import json

XLSX_FILE = 'HPV_25_26-zminy-hruden.xlsx'
OUTPUT_JSON = 'addresses_from_xlsx.json'

# Patterns
queue_pattern = re.compile(r'черг[аи]\s*(\d+)', re.IGNORECASE)
city_pattern = re.compile(r'(м\.|с\.|смт\.)\s*([А-ЯІЇЄҐ][а-яіїєґ\'\-]+)', re.IGNORECASE)

# Enterprise patterns to filter out
enterprise_patterns = [
    r'\bАТ\b', r'\bПАТ\b', r'\bТОВ\b', r'\bПрАТ\b', r'\bКП\b', r'\bПП\b',
    r'\bВАТ\b', r'\bДП\b', r'\bВО\b', r'\bФОП\b', r'\bТОВ\b',
    r'завод', r'комбінат', r'підприємств', r'компанія', r'філія',
    r'турбомеханічн', r'автоагрегатн', r'облагропостач'
]
enterprise_regex = re.compile('|'.join(enterprise_patterns), re.IGNORECASE)

# Street patterns - more flexible
street_markers = [
    'вул.', 'пров.', 'просп.', 'бульв.', 'пл.', 'майдан', 'узвіз'
]

def is_enterprise_address(text):
    """Check if text contains enterprise markers"""
    return bool(enterprise_regex.search(text))

def extract_addresses_from_text(text, current_queue):
    """Extract addresses from a text block with improved pattern matching"""
    if not text or not isinstance(text, str):
        return []

    addresses = []

    # Extract city if present
    cities = []
    for city_match in city_pattern.finditer(text):
        city = f"{city_match.group(1)} {city_match.group(2).strip()}"
        cities.append(city)
    default_city = cities[-1] if cities else "м. Полтава"

    # Remove enterprise mentions to avoid matching them
    clean_text = text
    for ent_pattern in [r'(?:АТ|ПАТ|ТОВ|ПрАТ|КП|ПП)\s+"[^"]+"', r'(?:АТ|ПАТ|ТОВ|ПрАТ|КП|ПП)\s+[А-ЯІЇЄҐ][^\.;,]{5,40}(?=[\.;,]|$)']:
        clean_text = re.sub(ent_pattern, ' ', clean_text, flags=re.IGNORECASE)

    # Split by period to get segments (each segment may contain multiple street-house pairs)
    # Periods usually separate different address groups
    segments = re.split(r'\.(?=\s)', clean_text)

    for segment in segments:
        segment = segment.strip()
        if len(segment) < 5:
            continue

        # Extract all street patterns from this segment
        # Pattern 1: Explicit markers like "вул. Name numbers"
        for marker in street_markers:
            # Match: marker + street name + all following house numbers until next street/marker
            pattern = rf'({marker})\s*([А-ЯІЇЄҐа-яіїєґ][А-ЯІЇЄҐа-яіїєґ\'\-\s]+?)[\s,]+((?:[0-9]+[а-яА-ЯіїєґІЇЄҐзЗ]?(?:/[0-9]+)?(?:\-[0-9]+[а-яА-ЯіїєґІЇЄҐзЗ]?)?[\s,]*)+)'

            for match in re.finditer(pattern, segment, re.IGNORECASE):
                street_type = match.group(1).strip()
                street_name = match.group(2).strip().rstrip(',').rstrip()
                houses_raw = match.group(3).strip()

                # Clean street name
                street_name = re.sub(r'\s+', ' ', street_name).strip()

                # Skip if street name looks like enterprise or is too short
                if is_enterprise_address(street_name) or len(street_name) < 2:
                    continue

                street_full = f"{street_type} {street_name}"

                # Parse all house numbers
                houses = parse_houses(houses_raw)

                for house in houses:
                    if house:  # Skip empty houses
                        addr = {
                            "city": default_city,
                            "street": street_full,
                            "house": house.strip(),
                            "queue": current_queue
                        }
                        if not is_enterprise_address(street_full):
                            addresses.append(addr)

        # Pattern 2: Implicit street names (StreetName, numbers, numbers, ...)
        # Match street name followed by comma and numbers
        # This handles: "Баленка, 10, 12, 14, 16" or "Баленка-10,12,14"
        implicit_pattern = r'(?:^|[,;\s])([А-ЯІЇЄҐ][а-яіїєґ\'\-]+(?:\s+[А-ЯІЇЄҐ][а-яіїєґ\'\-]+)?)[\s,\-]+((?:[0-9]+[а-яА-ЯіїєґІЇЄҐзЗ]?(?:/[0-9]+)?(?:\-[0-9]+[а-яА-ЯіїєґІЇЄҐзЗ]?)?[\s,]*)+)'

        for match in re.finditer(implicit_pattern, segment):
            street_name = match.group(1).strip()
            houses_raw = match.group(2).strip()

            # Skip if it looks like enterprise
            if is_enterprise_address(street_name):
                continue

            # Skip common words
            if street_name.lower() in ['полтава', 'київ', 'україна', 'буд', 'будинок', 'номер']:
                continue

            # Skip very short names or single letters
            if len(street_name) < 3:
                continue

            # Skip if it's likely not a street (e.g., starts with number)
            if street_name[0].isdigit():
                continue

            # Assume it's a street (add вул. prefix)
            street_full = f"вул. {street_name}"

            houses = parse_houses(houses_raw)
            for house in houses:
                if house:  # Skip empty houses
                    addr = {
                        "city": default_city,
                        "street": street_full,
                        "house": house.strip(),
                        "queue": current_queue
                    }
                    addresses.append(addr)

    return addresses

def parse_houses(houses_text):
    """Parse house numbers from text like '1,2,3-5,7а'"""
    houses = []

    # Clean up the input
    houses_text = houses_text.strip()

    # Remove trailing non-digit characters that might be part of enterprise names
    houses_text = re.sub(r'[а-яА-ЯіїєґІЇЄҐ]{3,}.*$', '', houses_text).strip()

    # Split by comma or semicolon
    parts = re.split(r'[,;]\s*', houses_text)

    for part in parts:
        part = part.strip()
        if not part or not re.search(r'\d', part):
            continue

        # Skip if it looks like enterprise code
        if len(part) > 20:
            continue

        # Handle ranges like "3-5" or "1-9"
        range_match = re.match(r'^(\d+)\s*-\s*(\d+)([а-яА-ЯіїєґІЇЄҐ]?)$', part)
        if range_match:
            start = int(range_match.group(1))
            end = int(range_match.group(2))
            suffix = range_match.group(3)
            if end - start <= 50:  # Reasonable range
                houses.extend([str(i) + suffix for i in range(start, end + 1)])
        else:
            # Single house number (can be like "3а", "12/5", etc)
            # Extract just the house number part
            house_match = re.match(r'^([0-9]+[а-яА-ЯіїєґІЇЄҐ]?(?:/[0-9]+)?)', part)
            if house_match:
                houses.append(house_match.group(1))

    return houses if houses else [""]

# Load workbook
wb = openpyxl.load_workbook(XLSX_FILE)

all_addresses = []
current_queue = None

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not any(cell for cell in row):
            continue

        # Combine all cells in row to process together
        row_text = ' '.join([str(cell) for cell in row if cell])

        # Check for queue number
        queue_match = queue_pattern.search(row_text)
        if queue_match:
            current_queue = int(queue_match.group(1))

        # Try to extract addresses if row contains street markers
        if any(marker in row_text.lower() for marker in street_markers):
            extracted = extract_addresses_from_text(row_text, current_queue)
            all_addresses.extend(extracted)

wb.close()

# Remove duplicates while preserving order
unique_addresses = []
seen = set()

for addr in all_addresses:
    # Create a key for uniqueness check
    key = (addr['city'], addr['street'], addr['house'], addr['queue'])
    if key not in seen:
        seen.add(key)
        unique_addresses.append(addr)

# Save to JSON
with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(unique_addresses, f, ensure_ascii=False, indent=2)

print(f"Total addresses extracted: {len(unique_addresses)}")
print(f"Unique cities: {len(set(a['city'] for a in unique_addresses))}")
print(f"Queues found: {sorted(set(a['queue'] for a in unique_addresses if a['queue']))}")
print(f"\nSample addresses:")
for addr in unique_addresses[:10]:
    print(f"  {addr['city']}, {addr['street']} {addr['house']} (Queue {addr['queue']})")
